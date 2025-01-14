import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelHelper:
    @classmethod
    def convert_html_to_excel(cls, html_file_path, output_excel_path):
        # Read and parse HTML
        with open(html_file_path, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file.read(), 'html.parser')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Water Budget"
        
        # Define styles
        header_fill = PatternFill(start_color="DEE2E6", end_color="DEE2E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add main heading "Jalagam"
        main_heading_row = 1
        for col in range(1, 14):  # Columns A through M (1-13)
            cell = ws.cell(row=main_heading_row, column=col)
            cell.border = Border(bottom=Side(style='medium'))
        ws.merge_cells(f'A{main_heading_row}:M{main_heading_row}')
        main_heading = ws.cell(row=main_heading_row, column=1, value="JALAGAM: The Water Budget Tool")
        main_heading.font = Font(size=24, bold=True)
        main_heading.alignment = Alignment(horizontal='center')

        
        # Set starting row for all sections
        section_start_row = 3
        
        # Add BASIC INFORMATION section heading
        ws.merge_cells(f'A{section_start_row}:B{section_start_row}')
        basic_header = ws.cell(row=section_start_row, column=1, value="BASIC INFORMATION")
        basic_header.font = Font(bold=True, size=14)
        basic_header.alignment = Alignment(horizontal='center')
        basic_header.fill = header_fill
        
        # Add DEMAND section heading
        col_offset_demand = 5
        ws.merge_cells(f'{get_column_letter(col_offset_demand)}{section_start_row}:{get_column_letter(col_offset_demand+2)}{section_start_row}')
        demand_header = ws.cell(row=section_start_row, column=col_offset_demand, value="DEMAND")
        demand_header.font = Font(bold=True, size=14)
        demand_header.alignment = Alignment(horizontal='center')
        demand_header.fill = header_fill
        
        # Add SUPPLY section heading
        col_offset_supply = 10
        ws.merge_cells(f'{get_column_letter(col_offset_supply)}{section_start_row}:{get_column_letter(col_offset_supply+2)}{section_start_row}')
        supply_header = ws.cell(row=section_start_row, column=col_offset_supply, value="SUPPLY")
        supply_header.font = Font(bold=True, size=14)
        supply_header.alignment = Alignment(horizontal='center')
        supply_header.fill = header_fill
        
        # Start content rows after headers
        current_row = section_start_row + 2
        
        # Function to write a table section
        def write_section(table, start_row):
            nonlocal current_row
            
            # Write table contents
            if table:
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all(['th', 'td'])
                    for col_idx, col in enumerate(cols, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=col.text.strip())
                        cell.border = border
                        if col.name == 'th' or 'fw-semibold' in col.get('class', []):
                            cell.font = Font(bold=True)
                        if 'text-end' in col.get('class', []):
                            cell.alignment = Alignment(horizontal='right')
                        if col.text.strip() in ['Demand', 'Supply','Budget']:
                            cell.alignment = Alignment(horizontal='center')
                            cell.font = Font(bold=True)
                        if 'action' in col.text.strip():
                            cell.font = Font(bold=True)
                    current_row += 1
                current_row += 1  # Add space between sections
        
        # Process Basic Information
        basic_info_table = soup.find('div', text='BASIC INFORMATION').find_next('table')
        write_section(basic_info_table, current_row)
        
        # Add WATER TRANSFER section heading
        ws.merge_cells(f'A{current_row}:B{current_row}')
        transfer_header = ws.cell(row=current_row, column=1, value="WATER TRANSFER")
        transfer_header.font = Font(bold=True, size=14)
        transfer_header.alignment = Alignment(horizontal='center')
        transfer_header.fill = header_fill
        current_row += 2
        
        # Process Water Transfer
        transfer_table = soup.find('div', text='WATER TRANSFER').find_next('table')
        write_section(transfer_table, current_row)
        
        # Add WATER BUDGET section heading
        ws.merge_cells(f'A{current_row}:B{current_row}')
        budget_header = ws.cell(row=current_row, column=1, value="WATER BUDGET")
        budget_header.font = Font(bold=True, size=14)
        budget_header.alignment = Alignment(horizontal='center')
        budget_header.fill = header_fill
        current_row += 2
        
        # Process Water Budget
        budget_table = soup.find('div', text='WATER BUDGET').find_next('table')
        write_section(budget_table, current_row)
        
        # Reset current_row for demand and supply sections
        current_row = section_start_row + 2
        
        # Function to write a section in a new column
        def write_section_new_column(section_title, table, col_start):
            nonlocal current_row
            
            # Write section header
            ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_start+2)}{current_row}')
            header_cell = ws.cell(row=current_row, column=col_start, value=section_title)
            # header_cell.fill = header_fill
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='left')
            current_row += 1

            # Custom message before table
            if not table and table_id == 'industry':
                current_row += 1
                ws.merge_cells(f'{get_column_letter(col_offset_demand)}{current_row}:{get_column_letter(col_offset_demand+2)}{current_row}')
                no_industry_cell = ws.cell(row=current_row, column=col_offset_demand, value="*There are no industries in this block.")
                no_industry_cell.font = Font(italic=True)
                no_industry_cell.alignment = Alignment(horizontal='center')
                current_row += 2
                
            elif table and table_id == 'human':
                coefficient_div = soup.find('div', {'id': 'coefficient'})
                if coefficient_div:
                    coefficient_text = coefficient_div.get_text(strip=True)
                    coefficient_text = ' '.join(coefficient_text.split())
                    ws.merge_cells(f'{get_column_letter(col_offset_demand)}{current_row}:{get_column_letter(col_offset_demand+2)}{current_row}')
                    coefficient_cell = ws.cell(row=current_row, column=col_offset_demand, value=coefficient_text)
                    coefficient_cell.font = Font(italic=True)
                    coefficient_cell.alignment = Alignment(horizontal='left')
                    current_row += 1
            
            # Write table contents
            if table:
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all(['th', 'td'])
                    for col_idx, col in enumerate(cols):
                        cell = ws.cell(row=current_row, column=col_start + col_idx, value=col.text.strip())
                        cell.border = border
                        if col.name == 'th' or 'fw-semibold' in col.get('class', []):
                            cell.font = Font(bold=True)
                        if 'text-end' in col.get('class', []):
                            cell.alignment = Alignment(horizontal='right')
                    current_row += 1
                current_row += 1
        
        # Process Demand sections
        demand_sections = [
            ('a) Human(in HaM)', 'human'),
            ('b) Livestock(in HaM)', 'livestock'),
            ('c) Crops(in HaM)', 'crops'),
            ('d) Industry(in HaM)', 'industry')
        ]
        
        for section_title, table_id in demand_sections:
            table = soup.find('table', {'id': table_id})
            write_section_new_column(section_title, table, col_offset_demand)
            
            
        # Reset current_row for supply sections
        current_row = section_start_row + 2
        
        # Process Supply sections
        supply_sections = [
            ('a) Surface Water(in HaM)', 'surface_water'),
            ('b) Groundwater(in HaM)', 'groundwater'),
            ('c) Runoff(in HaM)', 'runoff'),
            ('d) LULC(in HaM)', 'lulc'),
            ('e) Rainfall(in mm)', 'rainfall')
        ]
        
        for section_title, table_id in supply_sections:
            table = soup.find('table', {'id': table_id})
            write_section_new_column(section_title, table, col_offset_supply)
        
        # Adjust column widths
        for column in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(column)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=column)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_excel_path)
        print(f"Excel file created successfully at: {output_excel_path}")
        return output_excel_path