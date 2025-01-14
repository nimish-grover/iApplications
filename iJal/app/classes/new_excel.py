import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelGenerator:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Water Budget"
        
        # Define styles
        self.header_fill = PatternFill(start_color="DEE2E6", end_color="DEE2E6", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def create_water_budget_excel(self, data_dict, output_path):
        """
        Creates Excel file from dictionary data
        
        Args:
            data_dict (dict): Dictionary containing all section data
            output_path (str): Path to save Excel file
        """
        self._add_main_heading()
        current_row = 3
        
        # Add section headers
        self._add_section_headers(current_row)
        current_row += 1
        
        # Process Basic Information
        if 'basic_info' in data_dict:
            self._write_basic_info(data_dict['basic_info'], current_row)
            current_row = self.ws.max_row + 2
        
        # Process Water Transfer
        if 'transfer_data' in data_dict:
            self._write_section("WATER TRANSFER", self._format_transfer_data(data_dict['transfer_data']), current_row)
            current_row = self.ws.max_row + 2
        
        # Process Water Budget
        if 'water_budget' in data_dict:
            budget_data = {
                'demand_side': data_dict.get('demand_side', []),
                'supply_side': data_dict.get('supply_side', []),
                'water_budget': data_dict.get('water_budget', [])
            }
            self._write_section("WATER BUDGET", self._format_budget_data(budget_data), current_row)
        
        # Process Demand sections
        demand_row = 4
        col_offset_demand = 5
        
        # Add coefficient if present
        if 'coefficient' in data_dict:
            coeff_cell = self.ws.cell(row=demand_row, column=col_offset_demand, 
                                    value=f"*Water demand coefficient: {data_dict['coefficient']}L/day")
            coeff_cell.font = Font(italic=True)
            demand_row += 1
        
        demand_sections = [
            ('a) Human(in HaM)', 'human_data', self._format_human_data),
            ('b) Livestock(in HaM)', 'livestock_data', self._format_livestock_data),
            ('c) Crops(in HaM)', 'crop_data', self._format_crop_data),
            ('d) Industry(in HaM)', 'industry_data', self._format_industry_data)
        ]
        
        for title, key, formatter in demand_sections:
                
            if key in data_dict:
                formatted_data = formatter(data_dict[key])
                self._write_column_section(title, formatted_data, demand_row, col_offset_demand)
                demand_row = demand_row+len(formatted_data) + 2
                
        
        # Process Supply sections
        supply_row = 5
        col_offset_supply = 10
        supply_sections = [
            ('a) Surface Water(in HaM)', 'surface_water_data', self._format_surface_water_data),
            ('b) Groundwater(in HaM)', 'groundwater_data', self._format_groundwater_data),
            ('c) Runoff(in HaM)', 'runoff_data', self._format_runoff_data),
            ('d) LULC(in HaM)', 'lulc_data', self._format_lulc_data),
            ('e) Rainfall(in mm)', 'rainfall_data', self._format_rainfall_data)
        ]
        
        for title, key, formatter in supply_sections:
            if key in data_dict:
                formatted_data = formatter(data_dict[key])
                self._write_column_section(title, formatted_data, supply_row, col_offset_supply)
                supply_row = supply_row + len(formatted_data) + 2
        
        self._adjust_column_widths()
        self.wb.save(output_path)

    def _add_main_heading(self):
        """Adds the main Jalagam heading to the worksheet"""
        for col in range(1, 14):  # Columns A through M
            cell = self.ws.cell(row=1, column=col)
            cell.border = Border(bottom=Side(style='medium'))
            
        self.ws.merge_cells('A1:M1')
        main_heading = self.ws.cell(row=1, column=1, value="JALAGAM: The Water Budget Tool")
        main_heading.font = Font(size=24, bold=True)
        main_heading.alignment = Alignment(horizontal='center')

    def _add_section_headers(self, row):
        """Adds the main section headers (Basic Information, Demand, Supply)"""
        # Basic Information header
        self.ws.merge_cells(f'A{row}:B{row}')
        basic_header = self.ws.cell(row=row, column=1, value="BASIC INFORMATION")
        basic_header.font = Font(bold=True, size=14)
        basic_header.alignment = Alignment(horizontal='center')
        basic_header.fill = self.header_fill
        
        # Demand header
        self.ws.merge_cells(f'E{row}:G{row}')
        demand_header = self.ws.cell(row=row, column=5, value="DEMAND")
        demand_header.font = Font(bold=True, size=14)
        demand_header.alignment = Alignment(horizontal='center')
        demand_header.fill = self.header_fill
        
        # Supply header
        self.ws.merge_cells(f'J{row}:L{row}')
        supply_header = self.ws.cell(row=row, column=10, value="SUPPLY")
        supply_header.font = Font(bold=True, size=14)
        supply_header.alignment = Alignment(horizontal='center')
        supply_header.fill = self.header_fill

    def _write_basic_info(self, data, start_row):
        """Writes basic information section"""
        rows = []
        for key, value in data.items():
            rows.append([key, value])

        self._write_data_with_headers(rows, [], start_row)

    def _write_section(self, section_title, data, start_row):
        """Writes a full section with title and data"""
        # Write section header
        self.ws.merge_cells(f'A{start_row}:B{start_row}')
        header = self.ws.cell(row=start_row, column=1, value=section_title)
        header.font = Font(bold=True, size=14)
        header.alignment = Alignment(horizontal='center')
        header.fill = self.header_fill
        
        # Write data starting from next row
        for row_idx, row_data in enumerate(data, start=start_row + 2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
                # Style based on content
                if row_idx == start_row + 2:  # Header row
                    cell.font = Font(bold=True)
                elif row_data[0] in ['Demand', 'Supply', 'Budget']:  # Section headers
                    self.ws.merge_cells(f'A{row_idx}:B{row_idx}')
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    break  # Skip second column for merged cells
                elif col_idx == 2 and isinstance(value, (int, float)):  # Number values
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='right')
                
                # Style the action message row
                if "Water is" in str(row_data[0]):
                    cell.font = Font(bold=True)
                    if col_idx == 2:  # Value column
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')

    def _write_column_section(self, section_title, data, start_row, col_offset):
        """Writes a section in a specific column"""
        # Write section header
        self.ws.merge_cells(f'{get_column_letter(col_offset)}{start_row}:'
                          f'{get_column_letter(col_offset+2)}{start_row}')
        header = self.ws.cell(row=start_row, column=col_offset, value=section_title)
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal='left')
        
        # Write data
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            for col_idx, value in enumerate(row_data):
                # Skip cells that are part of a merged range (but not the top-left cell)
                target_cell = self.ws.cell(row=row_idx, column=col_offset + col_idx)
                if any(
                    target_cell.coordinate in merged_range
                    and target_cell.coordinate != merged_range.start_cell.coordinate
                    for merged_range in self.ws.merged_cells.ranges
                ):
                    continue
                
                # Write the value if it's not skipped
                target_cell.value = value
                target_cell.border = self.border
                if row_idx == start_row + 1:  # Header row
                    target_cell.font = Font(bold=True)
                if isinstance(value, (int, float)):
                    target_cell.alignment = Alignment(horizontal='right')

    def _write_data_with_headers(self, data, headers, start_row):
        """Helper method to write data with headers"""
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.border = self.border
        
        # Write data
        current_row = start_row + 1
        for row_data in data:
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.border
            current_row += 1

    def _format_human_data(self, data):
        """Format human data from array of dictionaries"""
        headers = ['Category', 'Population', 'Consumption']
        rows = [[str(item['category']).capitalize(), item.get('count', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_livestock_data(self, data):
        """Format livestock data from array of dictionaries"""
        headers = ['Type', 'Count', 'Consumption']
        rows = [[str(item['category']).capitalize(), item.get('count', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_crop_data(self, data):
        """Format crop data from array of dictionaries"""
        headers = ['Crop', 'Area', 'Consumption']
        rows = [[str(item['category']).capitalize(), item.get('count', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_industry_data(self, data):
        """Format industry data from array of dictionaries"""
        if not data:
            return [['*No industries in this block']]
        headers = ['Industry', 'Type', 'Water Allocation']
        rows = [[str(item['category']).capitalize(), item.get('count', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_surface_water_data(self, data):
        """Format surface water data from array of dictionaries"""
        headers = ['Water Body', 'Count','Storage Capacity']
        rows = [[item.get('category', ''), item.get('count', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_groundwater_data(self, data):
        """Format groundwater data from array of dictionaries"""
        headers = ['Description', 'Value']
        rows = [[item.get('name', ''), item.get('value', '')] for item in data]
        return [headers] + rows

    def _format_runoff_data(self, data):
        """Format water transfer data from array of dictionaries"""
        headers = ['Catchment', "% Runoff", 'Yield','Available']
        rows = [[item.get('catchment', ''), item.get('runoff', ''), item.get('runoff_yield', ''), item.get('supply', '')] for item in data]
        return [headers] + rows
    
    def _format_lulc_data(self, data):
        """Format lulc data from array of dictionaries"""
        headers = ['Name', 'Area']
        rows = [[item.get('lulc_name', ''), item.get('lulc_area', '')] for item in data]
        return [headers] + rows
    
    def _format_transfer_data(self, data):
        """Format water transfer data from array of dictionaries"""
        headers = ['Water Transfer', 'Quantity']
        rows = [[item.get('entity_name', ''), item.get('entity_value', '')] for item in data]
        return [headers] + rows
    

    def _format_rainfall_data(self, data):
        """Format rainfall data from array of dictionaries"""
        headers = ['Month-Year', 'Actual','Normal']
        rows = [[item.get('month', ''), item.get('actual', ''), item.get('normal', '')] for item in data]
        return [headers] + rows

    def _format_budget_data(self, data_dict):
        """
        Format water budget data matching the HTML structure
        
        Args:
            data_dict (dict): Dictionary containing demand_side, supply_side, and water_budget lists
        """
        headers = ['Description', 'Value']
        rows = []
        
        # Add Demand section
        rows.append(['Demand', ''])  # Demand header
        for item in data_dict.get('demand_side', []):
            rows.append([
                item['category'].title(),
                float(item['water_value'])
            ])
        
        # Add Supply section
        rows.append(['Supply', ''])  # Supply header
        for item in data_dict.get('supply_side', []):
            rows.append([
                item['category'].title(),
                float(item['water_value'])
            ])
        
        # Add Budget section
        rows.append(['Budget', ''])  # Budget header
        for item in data_dict.get('water_budget', []):
            rows.append([
                item['category'].title(),
                float(item['water_value'])
            ])
        
        # Calculate and add final budget row
        
        total_supply = float(data_dict['water_budget'][1]['water_value'])
        total_demand = float(data_dict['water_budget'][0]['water_value'])
        water_budget_value = total_supply - total_demand
        
        message = ("Water is Deficient in Block, Please take necessary actions." 
                if water_budget_value < 0 
                else "Water is Surplus in Block, No action is required.")
        
        rows.append([message, water_budget_value])
        
        return [headers] + rows

    def _adjust_column_widths(self):
        """Adjusts column widths based on content"""
        for column in range(1, self.ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(column)
            
            for row in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row, column=column)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column_letter].width = adjusted_width